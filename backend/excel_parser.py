import os
import re
from typing import List, Dict, Tuple, Optional
from openpyxl import load_workbook
from backend.models import (
    ReachData, DeviceBreakdown, CreativeBreakdown, 
    AgeBreakdown, GenderBreakdown, CampaignData, TemplateMetadata
)
from datetime import datetime

class ExcelParser:
    """Parses input Excel files and template"""
    
    REQUIRED_SHEETS = ['REACH', 'DATE', 'APP URL', 'TIME OF DAY', 'EXCHANGE', 
                       'DEVICE', 'CREATIVE', 'CITY', 'AGE', 'GENDER']
    
    @staticmethod
    def parse_filename(filename: str) -> Tuple[str, str]:
        """
        Try to extract audience and burst number from filename.
        Returns (audience, burst_number). Fallbacks to filename/1 if not found.
        """
        filename_lower = filename.lower()
        
        # Match audience (Try to find known ones, otherwise use filename)
        if 'vietnamese' in filename_lower:
            audience = 'Vietnamese'
        elif 'punjabi' in filename_lower:
            audience = 'Punjabi'
        elif 'filipino' in filename_lower:
            audience = 'Filipino'
        else:
            # Fallback: Use the first part of the filename as the audience
            audience = filename.split('_')[0].split('-')[0].split('.')[0]
        
        # Match burst number (Look for 'burst' followed by a number)
        burst_match = re.search(r'burst[_-]?(\d+)', filename_lower)
        if burst_match:
            burst_number = burst_match.group(1)
        else:
            # Fallback: Assume Burst 1
            burst_number = "1"
        
        return audience, burst_number

    @staticmethod
    def extract_audience_label(filename: str, selected_format: str = "") -> str:
        """Extract the report audience label from a filename without using the full file name."""
        stem = filename.rsplit('.', 1)[0]
        parts = [part for part in re.split(r'[_\s]+', stem) if part]
        format_value = (selected_format or "").strip().lower()

        if format_value:
            for index, part in enumerate(parts):
                if part.lower() == format_value:
                    return "_".join(parts[:index + 1])

        for index, part in enumerate(parts):
            if part.lower() in {"banner", "video"}:
                return "_".join(parts[:index + 1])

        for marker in ["_Final_Report", "_Report", "_Final"]:
            marker_index = stem.lower().find(marker.lower())
            if marker_index > 0:
                return stem[:marker_index]

        return stem

    @staticmethod
    def _extract_audience_from_creative(creatives: List[CreativeBreakdown]) -> Optional[str]:
        """Extract audience name (e.g., Filipino) from creative names"""
        if not creatives:
            return None
            
        for creative in creatives:
            name = creative.name
            # Common pattern: RYMES0055_CA01501_Filipino_...
            # We look for Vietnamese, Punjabi, Filipino, etc.
            name_lower = name.lower()
            if 'vietnamese' in name_lower: return 'Vietnamese'
            if 'punjabi' in name_lower: return 'Punjabi'
            if 'filipino' in name_lower: return 'Filipino'
            
            # Fallback: Try to take the 3rd part of the underscore-separated name
            parts = name.split('_')
            if len(parts) >= 3:
                # Usually: CODE_CODE_AUDIENCE_...
                return parts[2]
                
        return None

    @staticmethod
    def parse_input_file(filepath: str, selected_format: str = "") -> CampaignData:
        """Parse a single input campaign report file"""
        wb = load_workbook(filepath, data_only=True)
        
        # Only REACH is strictly required for core metrics
        if 'REACH' not in wb.sheetnames:
            raise ValueError("Missing required sheet: REACH")
        
        # Extract filename info as a baseline
        filename = os.path.basename(filepath)
        file_audience, burst_number = ExcelParser.parse_filename(filename)
        audience_label = ExcelParser.extract_audience_label(filename, selected_format)
        
        # Extract CREATIVE data FIRST to get the real audience name
        creative_breakdown = []
        if 'CREATIVE' in wb.sheetnames:
            creative_breakdown = ExcelParser._parse_creative_sheet(wb['CREATIVE'])
            
        # Try to get audience from creative sheet, fallback to filename
        extracted_audience = ExcelParser._extract_audience_from_creative(creative_breakdown)
        audience = extracted_audience if extracted_audience else file_audience
        
        # Extract REACH data
        reach_sheet = wb['REACH']
        reach_data = ExcelParser._parse_reach_sheet(reach_sheet)
        
        # Extract other optional sheets
        device_breakdown = []
        if 'DEVICE' in wb.sheetnames:
            device_breakdown = ExcelParser._parse_device_sheet(wb['DEVICE'])
            
        age_breakdown = []
        if 'AGE' in wb.sheetnames:
            age_breakdown = ExcelParser._parse_age_sheet(wb['AGE'])
        else:
            age_breakdown = [AgeBreakdown(age_band=b, impressions=0, clicks=0, ctr=0) for b in ['18-24', '25-34', '35-44', '45-54']]
            
        gender_breakdown = []
        if 'GENDER' in wb.sheetnames:
            gender_breakdown = ExcelParser._parse_gender_sheet(wb['GENDER'])
        else:
            gender_breakdown = [GenderBreakdown(gender=g, impressions=0, clicks=0, ctr=0) for g in ['Male', 'Female']]
            
        # Extract dates
        start_date, end_date = datetime.today(), datetime.today()
        complete_views = 0
        if 'DATE' in wb.sheetnames:
            start_date, end_date = ExcelParser._parse_date_sheet(wb['DATE'])
            complete_views = ExcelParser._parse_complete_views(wb['DATE'])
        
        return CampaignData(
            audience=audience,
            burst_number=burst_number,
            audience_label=audience_label,
            reach_data=reach_data,
            device_breakdown=device_breakdown,
            creative_breakdown=creative_breakdown,
            age_breakdown=age_breakdown,
            gender_breakdown=gender_breakdown,
            start_date=start_date,
            end_date=end_date,
            complete_views=complete_views
        )
    
    @staticmethod
    def _parse_reach_sheet(sheet) -> ReachData:
        """Extract data from REACH sheet, Row 2"""
        row = 2
        actual_impressions = sheet.cell(row, 1).value or 0
        link_clicks = sheet.cell(row, 2).value or 0
        ctr = sheet.cell(row, 3).value or 0
        reach = sheet.cell(row, 4).value or 0
        frequency = sheet.cell(row, 5).value or 3
        
        return ReachData(
            actual_impressions=float(actual_impressions),
            link_clicks=float(link_clicks),
            ctr=float(ctr),
            reach=float(reach),
            frequency=float(frequency)
        )
    
    @staticmethod
    def _parse_device_sheet(sheet) -> List[DeviceBreakdown]:
        """Extract device breakdown data"""
        devices = []
        
        # Map source names to output names
        device_mapping = {
            'Mobile': 'Mobile',
            'Smart Phone': 'Mobile',
            'Tablet': 'Tablet',
            'Desktop': 'Desktop'
        }
        
        row = 2
        while True:
            device_name = sheet.cell(row, 1).value
            if device_name is None:
                break
            
            if str(device_name).strip().lower() == 'grand total':
                row += 1
                continue
            
            # Map device name
            mapped_name = device_mapping.get(device_name, device_name)
            
            impressions = sheet.cell(row, 2).value or 0
            clicks = sheet.cell(row, 3).value or 0
            ctr = sheet.cell(row, 4).value or 0
            
            devices.append(DeviceBreakdown(
                device_type=mapped_name,
                impressions=float(impressions),
                clicks=float(clicks),
                ctr=float(ctr)
            ))
            row += 1
        
        return devices
    
    @staticmethod
    def _parse_creative_sheet(sheet) -> List[CreativeBreakdown]:
        """Extract creative breakdown data"""
        creatives = []
        
        row = 2
        while True:
            creative_name = sheet.cell(row, 1).value
            if creative_name is None:
                break
            
            if str(creative_name).strip().lower() == 'grand total':
                row += 1
                continue
            
            impressions = sheet.cell(row, 2).value or 0
            clicks = sheet.cell(row, 3).value or 0
            ctr = sheet.cell(row, 4).value or 0
            
            creatives.append(CreativeBreakdown(
                name=str(creative_name),
                impressions=float(impressions),
                clicks=float(clicks),
                ctr=float(ctr)
            ))
            row += 1
        
        return creatives
    
    @staticmethod
    def _parse_age_sheet(sheet) -> List[AgeBreakdown]:
        """Extract age band breakdown data"""
        ages = []
        age_order = ['18-24', '25-34', '35-44', '45-54']
        
        for age_band in age_order:
            row = 2
            found = False
            while True:
                age_name = sheet.cell(row, 1).value
                if age_name is None:
                    break
                
                if str(age_name).strip().lower() == 'grand total':
                    row += 1
                    continue
                
                if str(age_name).strip() == age_band:
                    impressions = sheet.cell(row, 2).value or 0
                    clicks = sheet.cell(row, 3).value or 0
                    ctr = sheet.cell(row, 4).value or 0
                    
                    ages.append(AgeBreakdown(
                        age_band=age_band,
                        impressions=float(impressions),
                        clicks=float(clicks),
                        ctr=float(ctr)
                    ))
                    found = True
                    break
                row += 1
            
            # Add even if not found (will have 0 values)
            if not found:
                ages.append(AgeBreakdown(
                    age_band=age_band,
                    impressions=0,
                    clicks=0,
                    ctr=0
                ))
        
        return ages
    
    @staticmethod
    def _parse_gender_sheet(sheet) -> List[GenderBreakdown]:
        """Extract gender breakdown data"""
        genders = []
        gender_order = ['Male', 'Female']
        
        for gender in gender_order:
            row = 2
            found = False
            while True:
                gender_name = sheet.cell(row, 1).value
                if gender_name is None:
                    break
                
                if str(gender_name).strip().lower() == 'grand total':
                    row += 1
                    continue
                
                if str(gender_name).strip() == gender:
                    impressions = sheet.cell(row, 2).value or 0
                    clicks = sheet.cell(row, 3).value or 0
                    ctr = sheet.cell(row, 4).value or 0
                    
                    genders.append(GenderBreakdown(
                        gender=gender,
                        impressions=float(impressions),
                        clicks=float(clicks),
                        ctr=float(ctr)
                    ))
                    found = True
                    break
                row += 1
            
            # Add even if not found
            if not found:
                genders.append(GenderBreakdown(
                    gender=gender,
                    impressions=0,
                    clicks=0,
                    ctr=0
                ))
        
        return genders
    
    @staticmethod
    def _parse_date_sheet(sheet) -> Tuple[datetime, datetime]:
        """Extract start and end dates from DATE sheet column A"""
        dates = []
        row = 2  # Skip header row
        
        while True:
            cell_value = sheet.cell(row, 1).value
            if cell_value is None:
                break
            
            # Skip "Grand Total" rows
            if str(cell_value).strip().lower() == "grand total":
                row += 1
                continue
            
            # Parse date
            if isinstance(cell_value, datetime):
                parsed_date = cell_value
            elif isinstance(cell_value, str):
                try:
                    parsed_date = datetime.strptime(cell_value.strip(), "%d %B, %Y")
                except ValueError:
                    row += 1
                    continue
            else:
                row += 1
                continue
            
            dates.append(parsed_date)
            row += 1
        
        if not dates:
            # Fallback if no dates found
            today = datetime.today()
            return today, today
        
        start_date = min(dates)
        end_date = max(dates)
        return start_date, end_date

    @staticmethod
    def _parse_complete_views(sheet) -> float:
        """Extract total complete views from a video DATE sheet when present."""
        complete_views_col = None
        for cell in sheet[1]:
            if cell.value is None:
                continue

            header = str(cell.value).strip().lower()
            if 'complete' in header and 'view' in header:
                complete_views_col = cell.column
                break

        if complete_views_col is None:
            return 0

        total = 0
        row = 2
        while True:
            label = sheet.cell(row, 1).value
            if label is None:
                break

            if str(label).strip().lower() != 'grand total':
                total += float(sheet.cell(row, complete_views_col).value or 0)

            row += 1

        return total
    
    @staticmethod
    def parse_template_file(filepath: str) -> TemplateMetadata:
        """Parse template file to extract metadata"""
        wb = load_workbook(filepath, data_only=True)
        
        if 'MPN & CPN Breakdown' not in wb.sheetnames:
            raise ValueError("Template must contain 'MPN & CPN Breakdown' sheet")
        
        sheet = wb['MPN & CPN Breakdown']
        
        # Extract template metadata from the sheet
        # This will be customized based on actual template structure
        # For now, return placeholder values
        return TemplateMetadata(
            platform='MPN',
            format_type='Banner',
            booked_impressions={},
            start_date={},
            end_date={},
            reporting_date={},
            audience_labels={}
        )
