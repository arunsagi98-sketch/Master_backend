import random
from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from backend.models import CampaignData, TemplateMetadata

class ReportGenerator:
    """Generates the consolidated Master Performance Report with proper formatting"""

    THIN = Side(style='thin')
    THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    BLUE_FILL = PatternFill('solid', fgColor='0000FF')
    YELLOW_FILL = PatternFill('solid', fgColor='FFFF00')
    WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')

    FONT_HEADER_BLUE = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    FONT_BOLD_11 = Font(name='Calibri', size=11, bold=True)
    FONT_REGULAR_10 = Font(name='Calibri', size=10)
    FONT_BOLD_10 = Font(name='Calibri', size=10, bold=True)

    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ALIGN_CENTER_NOWRAP = Alignment(horizontal='center', vertical='center')

    FREQUENCY = 3  # Constant frequency for reach calculation
    EXPECTED_ORDER = [
        ('Vietnamese', '1'),
        ('Punjabi', '1'),
        ('Vietnamese', '2'),
        ('Punjabi', '2')
    ]

    def __init__(
        self,
        campaigns: List[CampaignData],
        template_metadata: TemplateMetadata,
        selected_platform: str = "MPN",
        selected_format: str = "Banner",
    ):
        self.campaigns = self._sort_campaigns(campaigns)
        self.template = template_metadata
        self.selected_platform = selected_platform or "MPN"
        self.selected_format = selected_format or "Banner"
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'MPN & CPN Breakdown'
        self.current_row = 1

    def _sort_campaigns(self, campaigns: List[CampaignData]) -> List[CampaignData]:
        """Sort campaigns: Known audiences first, then others alphabetically, then by burst."""
        def sort_key(campaign):
            key = (campaign.audience, str(campaign.burst_number))
            
            # Known priority order
            priority_list = [
                ('Vietnamese', '1'),
                ('Punjabi', '1'),
                ('Filipino', '1'),
                ('Vietnamese', '2'),
                ('Punjabi', '2'),
                ('Filipino', '2')
            ]
            
            try:
                # If it's in the priority list, use its index
                return (0, priority_list.index(key), "")
            except ValueError:
                # Otherwise, sort by audience name then burst number
                return (1, campaign.audience, campaign.burst_number)

        return sorted(campaigns, key=sort_key)

    def _set_cell(
        self,
        row: int,
        col: int,
        value=None,
        fill=None,
        font=None,
        alignment=None,
        border=None,
        number_format=None,
    ):
        cell = self.ws.cell(row, col, value)
        if fill is not None:
            cell.fill = fill
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        if border is not None:
            cell.border = border
        if number_format is not None:
            cell.number_format = number_format
        return cell

    def generate(self) -> Workbook:
        """Generate the full report"""
        self._write_header()
        self._write_overview_section()
        self._write_performance_breakdowns()
        
        # Set column widths
        self.ws.column_dimensions['A'].width = 4.73
        self.ws.column_dimensions['B'].width = 55.09
        self.ws.column_dimensions['C'].width = 14.73
        self.ws.column_dimensions['D'].width = 27.09
        # Columns E-S use default width of 12.63
        
        # Set row heights
        self.ws.sheet_format.defaultRowHeight = 15.75
        self.ws.sheet_format.customHeight = True
        self.ws.row_dimensions[5].height = 28.5  # Overview header row
        
        return self.wb

    def _write_header(self):
        self._set_cell(
            2,
            2,
            'Please consolidate the audience breakdown into one sheet:)',
            font=self.FONT_REGULAR_10,
            alignment=self.ALIGN_CENTER,
        )
        self.current_row = 4

    def _write_section_title(self, row: int, title: str):
        self._set_cell(
            row,
            2,
            title,
            fill=self.BLUE_FILL,
            font=self.FONT_HEADER_BLUE,
            alignment=self.ALIGN_CENTER_WRAP,
        )

    def _write_overview_section(self):
        self._write_section_title(self.current_row, 'Overview')
        self.current_row += 1

        headers = [
            'Platform', 'Format', 'Audience', 'Reporting Date', 'Start Date',
            'End Date', 'Booked Impressions', 'Actual Impressions', 'Campaign Pacing',
            'Impression Pacing', 'Reach', 'Frequency', 'Link Click', 'CTR',
            'Complete Views', 'VCR', 'Amount Spent', 'Investment'
        ]

        for idx, header in enumerate(headers, start=2):
            self._set_cell(
                self.current_row,
                idx,
                header,
                fill=self.WHITE_FILL,
                font=self.FONT_BOLD_11,
                alignment=self.ALIGN_CENTER_WRAP,
                border=self.THIN_BORDER,
            )

        self.current_row += 1

        for campaign in self.campaigns:
            self._write_overview_row(campaign)

    def _write_overview_row(self, campaign: CampaignData):
        row = self.current_row
        audience_label = campaign.audience_label
        reach_data = campaign.reach_data
        booked_impressions = 0
        is_video = self.selected_format.strip().lower() == 'video'
        complete_views = campaign.complete_views if is_video else 0

        values = [
            self.selected_platform,
            self.selected_format,
            audience_label,
            datetime.now(),
            campaign.start_date,
            campaign.end_date,
            booked_impressions,
            reach_data.actual_impressions,
            0,
            f'=IFERROR(I{row}/H{row},0)',
            reach_data.reach,
            reach_data.frequency,
            reach_data.link_clicks,
            f'=N{row}/I{row}',
            complete_views,
            '-',
            f'=S{row}*K{row}',
            None,
        ]

        formats = [
            None,
            None,
            None,
            'd"-"mmm',
            'd"-"mmm',
            'd"-"mmm',
            '#,##0',
            '#,##0',
            '0%',
            '0%',
            '#,##0',
            '#,##0',
            '#,##0',
            '0.00%',
            '#,##0',
            None,
            '$#,##0.00;[Red]\\-"$"#,##0.00',
            '$#,##0.00;[Red]\\-"$"#,##0.00',
        ]

        fills = [
            None,
            None,
            None,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
            None,
            None,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
            None,
            self.WHITE_FILL,
            None,
            None,
            None,
        ]

        fonts = [
            self.FONT_BOLD_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
            self.FONT_REGULAR_10,
        ]

        alignments = [
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_WRAP,
            self.ALIGN_CENTER_WRAP,
            self.ALIGN_CENTER_WRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
            self.ALIGN_CENTER_NOWRAP,
        ]

        for idx, value in enumerate(values, start=2):
            self._set_cell(
                row,
                idx,
                value,
                fill=fills[idx - 2],
                font=fonts[idx - 2],
                alignment=alignments[idx - 2],
                border=self.THIN_BORDER,
                number_format=formats[idx - 2],
            )

        self.current_row += 1

    def _write_performance_breakdowns(self):
        for campaign in self.campaigns:
            self._write_performance_section(campaign)

    def _write_performance_section(self, campaign: CampaignData):
        self.current_row += 2
        title = f'Performance breakdown - by Audience - {campaign.audience_label}'
        self._write_section_title(self.current_row, title)
        self.current_row += 1

        self._write_subsection_header('By Device')
        self._write_device_subsection(campaign)
        self.current_row += 2

        self._write_subsection_header('By Creative')
        self._write_creative_subsection(campaign)
        self.current_row += 2

        self._write_subsection_header('By Age')
        self._write_age_subsection(campaign)
        self.current_row += 2

        self._write_subsection_header('By Gender')
        self._write_gender_subsection(campaign)

    def _write_subsection_header(self, label: str):
        headers = [
            'Actual Impressions', 'Reach', 'Frequency',
            'Complete Views', 'Link Click', 'CTR', 'Amount Spent'
        ]

        self._set_cell(
            self.current_row,
            2,
            label,
            fill=self.YELLOW_FILL,
            font=self.FONT_BOLD_11,
            alignment=self.ALIGN_CENTER_NOWRAP,
        )

        for idx, header in enumerate(headers, start=3):
            self._set_cell(
                self.current_row,
                idx,
                header,
                fill=self.WHITE_FILL,
                font=self.FONT_BOLD_11,
                alignment=self.ALIGN_CENTER_WRAP,
                border=self.THIN_BORDER,
            )

        self.current_row += 1

    def _allocate_metric(self, total: float, impressions_list: List[float], seed: str) -> List[int]:
        """Allocate an overview total across breakdown rows with stable random variation."""
        total = int(round(total or 0))
        if not impressions_list or total <= 0:
            return [0] * len(impressions_list)

        rnd = random.Random(seed)
        weighted_impressions = []
        for impressions in impressions_list:
            noise = rnd.uniform(0, 0.1 * (impressions or 1))
            weighted_impressions.append(impressions + noise)

        total_weight = sum(weighted_impressions)
        if total_weight == 0:
            return [0] * len(impressions_list)

        raw_allocations = [total * (w / total_weight) for w in weighted_impressions]
        allocated = [int(round(value)) for value in raw_allocations]

        difference = total - sum(allocated)
        if difference != 0:
            index = max(range(len(allocated)), key=lambda i: allocated[i])
            allocated[index] += difference

        return allocated

    def _allocate_reach(self, total_reach: int, impressions_list: List[float], seed: str) -> List[int]:
        """Allocate overview reach across breakdown rows."""
        return self._allocate_metric(total_reach, impressions_list, seed)

    def _allocate_complete_views(self, campaign: CampaignData, impressions_list: List[float], seed: str) -> List[int]:
        """Allocate overview complete views across breakdown rows."""
        if not campaign.complete_views or campaign.complete_views <= 0:
            return [0] * len(impressions_list)
        return self._allocate_metric(campaign.complete_views, impressions_list, seed)

    def _write_breakdown_row(
        self,
        label: str,
        impressions: float,
        clicks: float,
        reach: int,
        complete_views: float,
        is_amount_dash: bool = False,
    ):
        row = self.current_row

        values = [
            label,
            impressions,
            reach,
            self.FREQUENCY,
            complete_views,
            clicks,
            f'=G{row}/C{row}',
            '-' if is_amount_dash else None,
        ]

        formats = [
            None,
            '#,##0',
            '#,##0',
            '#,##0',
            '#,##0',
            '#,##0',
            '0.00%',
            '$#,##0.00;[Red]\\-"$"#,##0.00',
        ]

        fills = [
            None,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
            self.WHITE_FILL,
        ]

        for idx, value in enumerate(values, start=2):
            self._set_cell(
                row,
                idx,
                value,
                fill=fills[idx - 2],
                font=self.FONT_REGULAR_10,
                alignment=self.ALIGN_CENTER_NOWRAP,
                border=self.THIN_BORDER,
                number_format=formats[idx - 2],
            )

        self.current_row += 1

    def _write_device_subsection(self, campaign: CampaignData):
        impressions = [device.impressions for device in campaign.device_breakdown]
        reaches = self._allocate_reach(campaign.reach_data.reach, impressions, f'{campaign.audience}-device')
        complete_views = self._allocate_complete_views(campaign, impressions, f'{campaign.audience}-device-complete-views')
        for device, reach, views in zip(campaign.device_breakdown, reaches, complete_views):
            self._write_breakdown_row(device.device_type, device.impressions, device.clicks, reach, views, is_amount_dash=True)

    def _write_creative_subsection(self, campaign: CampaignData):
        impressions = [creative.impressions for creative in campaign.creative_breakdown]
        reaches = self._allocate_reach(campaign.reach_data.reach, impressions, f'{campaign.audience}-creative')
        complete_views = self._allocate_complete_views(campaign, impressions, f'{campaign.audience}-creative-complete-views')
        for creative, reach, views in zip(campaign.creative_breakdown, reaches, complete_views):
            self._write_breakdown_row(creative.name, creative.impressions, creative.clicks, reach, views, is_amount_dash=True)

    def _write_age_subsection(self, campaign: CampaignData):
        impressions = [age.impressions for age in campaign.age_breakdown]
        reaches = self._allocate_reach(campaign.reach_data.reach, impressions, f'{campaign.audience}-age')
        complete_views = self._allocate_complete_views(campaign, impressions, f'{campaign.audience}-age-complete-views')
        for age, reach, views in zip(campaign.age_breakdown, reaches, complete_views):
            self._write_breakdown_row(age.age_band, age.impressions, age.clicks, reach, views, is_amount_dash=False)

    def _write_gender_subsection(self, campaign: CampaignData):
        impressions = [gender.impressions for gender in campaign.gender_breakdown]
        reaches = self._allocate_reach(campaign.reach_data.reach, impressions, f'{campaign.audience}-gender')
        complete_views = self._allocate_complete_views(campaign, impressions, f'{campaign.audience}-gender-complete-views')
        for gender, reach, views in zip(campaign.gender_breakdown, reaches, complete_views):
            self._write_breakdown_row(gender.gender, gender.impressions, gender.clicks, reach, views, is_amount_dash=False)
