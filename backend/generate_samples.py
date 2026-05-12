#!/usr/bin/env python3
"""
Utility script to generate sample test files for the MPN Campaign Report Generator
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

def create_sample_campaign_file(filename, audience, burst_number):
    """Create a sample campaign report file"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create required sheets
    sheets = ['REACH', 'DATE', 'APP URL', 'TIME OF DAY', 'EXCHANGE', 
              'DEVICE', 'CREATIVE', 'CITY', 'AGE', 'GENDER']
    
    for sheet_name in sheets:
        ws = wb.create_sheet(sheet_name)
        
        if sheet_name == 'REACH':
            # REACH: Row 2 has Impressions, Clicks, CTR, Reach, Frequency
            ws['A1'] = 'Impressions'
            ws['B1'] = 'Clicks'
            ws['C1'] = 'CTR'
            ws['D1'] = 'Reach'
            ws['E1'] = 'Frequency'
            
            ws['A2'] = 100000 + (int(audience == 'Punjabi') * 50000)
            ws['B2'] = 500 + (int(audience == 'Punjabi') * 200)
            ws['C2'] = 0.005
            ws['D2'] = 50000 + (int(audience == 'Punjabi') * 25000)
            ws['E2'] = 3
        
        elif sheet_name == 'DEVICE':
            # DEVICE: Device type, Impressions, Clicks, CTR
            ws['A1'] = 'Device'
            ws['B1'] = 'Impressions'
            ws['C1'] = 'Clicks'
            ws['D1'] = 'CTR'
            
            devices = [
                ('Mobile', 60000, 350, 0.0058),
                ('Tablet', 25000, 100, 0.0040),
                ('Desktop', 15000, 50, 0.0033)
            ]
            
            for idx, (device, impr, clicks, ctr) in enumerate(devices, start=2):
                ws.cell(idx, 1, device)
                ws.cell(idx, 2, impr)
                ws.cell(idx, 3, clicks)
                ws.cell(idx, 4, ctr)
        
        elif sheet_name == 'CREATIVE':
            # CREATIVE: Creative name, Impressions, Clicks, CTR
            ws['A1'] = 'Creative'
            ws['B1'] = 'Impressions'
            ws['C1'] = 'Clicks'
            ws['D1'] = 'CTR'
            
            creatives = [
                ('Banner 300x250 V1', 50000, 300, 0.006),
                ('Banner 728x90 V1', 35000, 150, 0.0043),
                ('Banner 300x600 V1', 15000, 50, 0.0033)
            ]
            
            for idx, (creative, impr, clicks, ctr) in enumerate(creatives, start=2):
                ws.cell(idx, 1, creative)
                ws.cell(idx, 2, impr)
                ws.cell(idx, 3, clicks)
                ws.cell(idx, 4, ctr)
        
        elif sheet_name == 'AGE':
            # AGE: Age band, Impressions, Clicks, CTR
            ws['A1'] = 'Age'
            ws['B1'] = 'Impressions'
            ws['C1'] = 'Clicks'
            ws['D1'] = 'CTR'
            
            age_bands = [
                ('18-24', 15000, 110, 0.0073),
                ('25-34', 25000, 150, 0.006),
                ('35-44', 20000, 100, 0.005),
                ('45-54', 18000, 80, 0.0044),
                ('55-64', 15000, 30, 0.002),
                ('65+', 7000, 15, 0.0021)
            ]
            
            for idx, (age, impr, clicks, ctr) in enumerate(age_bands, start=2):
                ws.cell(idx, 1, age)
                ws.cell(idx, 2, impr)
                ws.cell(idx, 3, clicks)
                ws.cell(idx, 4, ctr)
        
        elif sheet_name == 'GENDER':
            # GENDER: Gender, Impressions, Clicks, CTR
            ws['A1'] = 'Gender'
            ws['B1'] = 'Impressions'
            ws['C1'] = 'Clicks'
            ws['D1'] = 'CTR'
            
            genders = [
                ('Male', 60000, 350, 0.0058),
                ('Female', 40000, 150, 0.0038)
            ]
            
            for idx, (gender, impr, clicks, ctr) in enumerate(genders, start=2):
                ws.cell(idx, 1, gender)
                ws.cell(idx, 2, impr)
                ws.cell(idx, 3, clicks)
                ws.cell(idx, 4, ctr)
        
        else:
            # Other sheets: just add a placeholder
            ws['A1'] = sheet_name
    
    wb.save(filename)
    print(f"✓ Created: {filename}")


def create_sample_template_file(filename):
    """Create a sample template file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "MPN & CPN Breakdown"
    
    # Add some basic structure
    ws['A1'] = "MPN & CPN Breakdown Template"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A3'] = "This is a template for the consolidated report"
    ws['A4'] = "The application will populate this sheet with campaign data"
    
    # Add headers for the overview section (optional pre-structure)
    headers = [
        'Platform', 'Format', 'Audience', 'Reporting Date', 'Start Date', 
        'End Date', 'Booked Impressions', 'Actual Impressions', 'Campaign Pacing',
        'Impression Pacing', 'Reach', 'Frequency', 'Link Click', 'CTR',
        'Complete Views', 'VCR', 'Amount Spent', 'Investment'
    ]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(6, col_idx, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
    
    wb.save(filename)
    print(f"✓ Created: {filename}")


def main():
    """Generate sample test files"""
    print("MPN Campaign Report Generator - Sample File Generator")
    print("=" * 50)
    
    # Create output directory
    sample_dir = "sample_files"
    os.makedirs(sample_dir, exist_ok=True)
    
    print(f"\nGenerating sample files in '{sample_dir}/' directory...")
    print()
    
    # Generate campaign files
    campaigns = [
        ('Vietnamese', '1'),
        ('Punjabi', '1'),
        ('Vietnamese', '2'),
        ('Punjabi', '2')
    ]
    
    for audience, burst in campaigns:
        filename = os.path.join(
            sample_dir, 
            f"Report_{audience}_Burst-{burst}_Final.xlsx"
        )
        create_sample_campaign_file(filename, audience, burst)
    
    # Generate template file
    template_filename = os.path.join(sample_dir, "Report_Template.xlsx")
    create_sample_template_file(template_filename)
    
    print()
    print("=" * 50)
    print("✓ Sample files created successfully!")
    print()
    print("Next steps:")
    print("1. Run: python run.py")
    print("2. Open http://localhost:8000 in your browser")
    print(f"3. Upload files from the '{sample_dir}/' directory")
    print("4. Click 'Generate Report' to create the consolidated report")
    print()


if __name__ == "__main__":
    main()
